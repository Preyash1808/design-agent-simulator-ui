"""
Routes for run-related operations.
Handles report generation, preprocessing, and test execution.
"""
import os
import pathlib
import json
import asyncio
import traceback
import uuid
import time
import re
import random
import string
import subprocess
import io
import csv
import zipfile
from typing import Dict, Any, Optional, List
from fastapi import APIRouter, Header, HTTPException, BackgroundTasks, UploadFile, File, Form
from fastapi.responses import FileResponse, Response

from ..storage import (
    get_supabase,
    use_supabase_db,
    upload_log_to_supabase,
    upload_file_to_supabase,
    upload_run_artifacts,
)
from ..utils import write_json
from ..models import PreprocessReq, TestsReq
from ..auth_utils import get_current_user
from ..db import fetchrow, execute
from ..report_builder import build_report_pdf, set_runs_path
from ..ingest import _ingest_run_artifacts
from ..persona_matcher import resolve_personas
from ..metrics import get_run_metrics_public
from ..storage import use_supabase_db, get_supabase

# Will be set by main.py
ROOT = None
RUNS = None
PYTHON = None


def set_paths(root: pathlib.Path, runs: pathlib.Path, python: str):
    """Set paths from main.py"""
    global ROOT, RUNS, PYTHON
    ROOT = root
    RUNS = runs
    PYTHON = python


router = APIRouter()


@router.get('/runs/{run_id}/report.pdf')
async def generate_pdf_report(run_id: str):
    try:
        data = await get_run_metrics_public(run_id)
    except Exception:
        raise HTTPException(status_code=404, detail='run not found')
    pdf_bytes = build_report_pdf(data, run_id)
    headers = {
        'Content-Disposition': f'attachment; filename="report_{run_id}.pdf"',
        'Content-Length': str(len(pdf_bytes)),
        'Cache-Control': 'no-store',
    }
    return Response(content=pdf_bytes, media_type='application/pdf', headers=headers)

@router.get('/runs/{run_id}/personas')
async def personas_summary(run_id: str):
    """Return per‑persona summary metrics to power Persona cards.

    Fields per persona: persona_id, name (if available), steps, completed, backtracks,
    waits, dropoffs_count, friction_pct (approx), sentiment_start/end (if available).
    """
    # Check if run exists (either locally or in database)
    run_dir = RUNS / run_id
    run_exists_locally = run_dir.exists()
    
    # If not local, check database
    if not run_exists_locally and use_supabase_db():
        try:
            client = get_supabase()
            # Check if run exists in database
            rr = client.table('run_results').select('run_id').eq('run_id', run_id).limit(1).execute()
            if not rr.data:
                raise HTTPException(status_code=404, detail=f'run not found in database: {run_id}')
        except Exception as e:
            print(f"Database query error for run {run_id}: {e}")
            raise HTTPException(status_code=500, detail=f'database error: {str(e)}')
    elif not run_exists_locally:
        raise HTTPException(status_code=404, detail='run not found')
    
    out: List[Dict[str, Any]] = []
    # Prefer Supabase run_results + optional insights
    try:
        if use_supabase_db():
            client = get_supabase()
            # Map filesystem run id to DB id if necessary
            db_run_id = run_id
            try:
                r0 = client.table('runs').select('id,run_dir').eq('id', run_id).limit(1).execute()
                if r0.data:
                    db_run_id = r0.data[0]['id']
                else:
                    abs_path = str(RUNS / run_id)
                    r1 = client.table('runs').select('id,run_dir').eq('run_dir', abs_path).limit(1).execute()
                    if r1.data:
                        db_run_id = r1.data[0]['id']
                    else:
                        # fallback: match by suffix '/<run_id>'
                        r2 = client.table('runs').select('id,run_dir').like('run_dir', f"%/{run_id}").limit(1).execute()
                        if r2.data:
                            db_run_id = r2.data[0]['id']
            except Exception:
                db_run_id = run_id
            rr = client.table('run_results').select('*').eq('run_id', db_run_id).execute()
            rows = rr.data or []
            # Optional: persona summaries from persona_summary.json in storage not readily accessible; keep DB-only for now
            for r in rows:
                out.append({
                    'persona_id': str(r.get('persona_id') or ''),
                    'steps': int(r.get('steps') or 0),
                    'completed': str(r.get('status') or '').lower() == 'completed',
                    'backtracks': int(r.get('backtracks') or 0) if 'backtracks' in r else None,
                    'wait_sec': float(r.get('time_sec') or 0.0),
                    'dropoff_reason': r.get('dropoff_reason') or None,
                })
    except Exception:
        out = []

    # Remove local file fallback: DB only per request

    # Aggregate card metrics
    cards: Dict[str, Dict[str, Any]] = {}
    for r in out:
        pid = str(r.get('persona_id') or '')
        c = cards.setdefault(pid, {'persona_id': pid, 'steps': [], 'completed': 0, 'runs': 0, 'backtracks': 0, 'dropoffs': 0, 'friction_count': 0})
        c['runs'] += 1
        c['steps'].append(int(r.get('steps') or 0))
        if r.get('completed'):
            c['completed'] += 1
        bt = r.get('backtracks')
        c['backtracks'] += int(bt or 0)
        # Count true exits only: increment when the session did NOT complete
        if not r.get('completed'):
            c['dropoffs'] += 1
    # Load persona names priority: runs.meta.persona_(plan|resolution) -> run_persona.name
    persona_names: Dict[str, str] = {}
    try:
        if use_supabase_db():
            try:
                client = get_supabase()
                db_run_id_local = run_id
                try:
                    r0 = client.table('runs').select('id').eq('id', run_id).limit(1).execute()
                    if r0.data:
                        db_run_id_local = r0.data[0]['id']
                except Exception:
                    db_run_id_local = run_id
                # 1) names from runs.meta first (persona_plan preferred, then persona_resolution)
                rrmeta = client.table('runs').select('meta').eq('id', db_run_id_local).limit(1).execute()
                if rrmeta.data:
                    meta = rrmeta.data[0].get('meta') or {}
                    for key in ['persona_plan', 'persona_resolution']:
                        try:
                            plist = (meta.get(key) or {}).get('personas') or []
                            for pr in plist:
                                pid = str(pr.get('slot') or pr.get('id') or '')
                                nm = str(pr.get('name') or '')
                                if pid and nm and pid not in persona_names:
                                    persona_names[pid] = nm
                        except Exception:
                            continue
                # 2) Do NOT fallback to run_persona names to avoid leaking user names.
                #    If meta has no name for a slot, we'll render a generic label later.
            except Exception:
                pass
        # Intentionally avoid static fallback from persona.json because those IDs are user IDs,
        # not persona slots; default naming handled later as 'Persona {slot}'.
    except Exception:
        persona_names = {}

    # Drift + sentiments per persona from TEA (database)
    drift_map: Dict[str, float] = {}
    sentiments_map: Dict[str, Dict[str, float]] = {}
    try:
        if use_supabase_db():
            try:
                client = get_supabase()
                # Map run id (again here to keep scopes independent of earlier blocks)
                db_run_id_local = run_id
                try:
                    r0 = client.table('runs').select('id,run_dir').eq('id', run_id).limit(1).execute()
                    if r0.data:
                        db_run_id_local = r0.data[0]['id']
                    else:
                        abs_path = str(RUNS / run_id)
                        r1 = client.table('runs').select('id,run_dir').eq('run_dir', abs_path).limit(1).execute()
                        if r1.data:
                            db_run_id_local = r1.data[0]['id']
                except Exception:
                    db_run_id_local = run_id
                tt = client.table('run_persona_teas').select('persona_id,sentiment_start,sentiment_end').eq('run_id', db_run_id_local).execute()
                rows = tt.data or []
                for row in rows:
                    try:
                        pid = str(row.get('persona_id') or '')
                        s0 = row.get('sentiment_start')
                        s1 = row.get('sentiment_end')
                        if isinstance(s0, (int, float)) and isinstance(s1, (int, float)):
                            drift_map[pid] = round(float(s1) - float(s0), 2)
                            sentiments_map[pid] = { 'start': float(s0), 'end': float(s1) }
                    except Exception:
                        continue
            except Exception:
                drift_map = {}
                sentiments_map = {}
    except Exception:
        drift_map = {}
        sentiments_map = {}

    # Get friction points count per persona
    friction_counts: Dict[str, int] = {}
    try:
        if use_supabase_db():
            try:
                client = get_supabase()
                db_run_id_local = run_id
                try:
                    r0 = client.table('runs').select('id,run_dir').eq('id', run_id).limit(1).execute()
                    if r0.data:
                        db_run_id_local = r0.data[0]['id']
                    else:
                        abs_path = str(RUNS / run_id)
                        r1 = client.table('runs').select('id,run_dir').eq('run_dir', abs_path).limit(1).execute()
                        if r1.data:
                            db_run_id_local = r1.data[0]['id']
                except Exception:
                    db_run_id_local = run_id
                # Get friction points count per persona
                fp = client.table('friction_points').select('persona_id').eq('run_id', db_run_id_local).execute()
                friction_rows = fp.data or []
                for fr in friction_rows:
                    pid = str(fr.get('persona_id') or '')
                    friction_counts[pid] = friction_counts.get(pid, 0) + 1
            except Exception:
                friction_counts = {}
    except Exception:
        friction_counts = {}

    # Compute derived
    items: List[Dict[str, Any]] = []
    for pid, c in cards.items():
        avg_steps = round(sum(c['steps'])/float(len(c['steps']) or 1), 1)
        completion_pct = round(100.0 * float(c['completed']) / float(c['runs'] or 1), 1)
        # Use actual friction points count instead of backtracks
        friction_count = friction_counts.get(pid, 0)
        friction_pct = round(100.0 * (float(friction_count)/float(sum(c['steps']) or 1)), 1) if c['steps'] else 0.0
        # Final persona display name: prefer DB/meta names; fallback to generic slot label
        persona_name = persona_names.get(pid) or f'Persona {pid}'
        items.append({
            'persona_id': pid,
            'persona_name': persona_name,
            'avg_steps': avg_steps,
            'completion_pct': completion_pct,
            'dropoffs': int(c['dropoffs']),
            'friction_pct': friction_pct,
            'drift': drift_map.get(pid),
            'sentiment_start': (sentiments_map.get(pid) or {}).get('start'),
            'sentiment_end': (sentiments_map.get(pid) or {}).get('end'),
        })
    return {'personas': items}

async def _analyze_unique_teas(run_id: str, persona_id: str) -> List[Dict[str, Any]]:
    """Analyze Unique TEAs (outliers) to spot edge cases and unusual user behaviors.
    
    Returns a list of outlier patterns that could indicate UX issues:
    - Users with unusually high step counts
    - Users who got stuck in loops
    - Users with extreme emotion patterns
    - Users who took very different paths
    - Users with unusual completion times
    """
    outliers = []
    
    try:
        if use_supabase_db():
            client = get_supabase()
            # Map run id
            db_run_id = run_id
            try:
                r0 = client.table('runs').select('id,run_dir').eq('id', run_id).limit(1).execute()
                if r0.data:
                    db_run_id = r0.data[0]['id']
                else:
                    abs_path = str(RUNS / run_id)
                    r1 = client.table('runs').select('id,run_dir').eq('run_dir', abs_path).limit(1).execute()
                    if r1.data:
                        db_run_id = r1.data[0]['id']
            except Exception:
                db_run_id = run_id
            
            # Get all results for this persona
            rr = client.table('run_results').select('*').eq('run_id', db_run_id).eq('persona_id', persona_id).execute()
            results = rr.data or []
            
            if not results:
                return outliers
            
            # Calculate statistics for outlier detection
            steps = [int(r.get('steps', 0)) for r in results]
            times = [float(r.get('time_sec', 0)) for r in results]
            completed = [r.get('status') == 'completed' for r in results]
            
            if not steps:
                return outliers
            
            # Calculate thresholds for outliers
            avg_steps = sum(steps) / len(steps)
            std_steps = (sum((s - avg_steps) ** 2 for s in steps) / len(steps)) ** 0.5
            high_step_threshold = avg_steps + 2 * std_steps
            
            avg_time = sum(times) / len(times)
            std_time = (sum((t - avg_time) ** 2 for t in times) / len(times)) ** 0.5
            high_time_threshold = avg_time + 2 * std_time
            
            # Find outliers
            for i, result in enumerate(results):
                step_count = int(result.get('steps', 0))
                time_sec = float(result.get('time_sec', 0))
                status = result.get('status', '')
                user_id = result.get('user_id', '')
                
                # High step count outlier
                if step_count > high_step_threshold:
                    outliers.append({
                        'type': 'high_steps',
                        'description': f'User took {step_count} steps (avg: {avg_steps:.1f})',
                        'severity': 'high' if step_count > avg_steps + 3 * std_steps else 'medium',
                        'user_id': user_id,
                        'value': step_count,
                        'threshold': high_step_threshold
                    })
                
                # Long completion time outlier
                if time_sec > high_time_threshold and status == 'completed':
                    outliers.append({
                        'type': 'long_time',
                        'description': f'User took {time_sec:.1f}s to complete (avg: {avg_time:.1f}s)',
                        'severity': 'high' if time_sec > avg_time + 3 * std_time else 'medium',
                        'user_id': user_id,
                        'value': time_sec,
                        'threshold': high_time_threshold
                    })
                
                # Incomplete session outlier (if most others completed)
                completion_rate = sum(completed) / len(completed)
                if status != 'completed' and completion_rate > 0.7:
                    outliers.append({
                        'type': 'incomplete',
                        'description': f'User did not complete (completion rate: {completion_rate:.1%})',
                        'severity': 'medium',
                        'user_id': user_id,
                        'value': status,
                        'threshold': 'completed'
                    })
                
                # Very short session outlier (potential immediate abandonment)
                if step_count < 3 and status != 'completed':
                    outliers.append({
                        'type': 'immediate_abandonment',
                        'description': f'User abandoned after only {step_count} steps',
                        'severity': 'high',
                        'user_id': user_id,
                        'value': step_count,
                        'threshold': 3
                    })
            
            # Sort by severity and limit to top outliers
            severity_order = {'high': 3, 'medium': 2, 'low': 1}
            outliers.sort(key=lambda x: (severity_order.get(x['severity'], 0), x['value']), reverse=True)
            
    except Exception as e:
        print(f"Error analyzing unique TEAs: {e}")
    
    return outliers[:10]  # Return top 10 outliers

@router.get('/runs/{run_id}/persona/{persona_id}')
async def persona_detail(run_id: str, persona_id: str):
    """Return per‑persona details for a run: TEA summary, path distribution, and exits/backtracks.

    Data sources (best-effort):
    - Supabase tables: run_persona_teas, run_dropoffs (optionally filtered by persona_id)
    - llm_run_insights.persona_teas JSON (fallback)
    - Local artifacts under runs/<id>/tests/persona_<persona_id>/simulations/*/{path.json,user_report.json}
    - Aggregated persona_summary.json
    """
    # Check if run exists (either locally or in database)
    run_dir = RUNS / run_id
    run_exists_locally = run_dir.exists()
    
    # If not local, check database
    if not run_exists_locally and use_supabase_db():
        try:
            client = get_supabase()
            # Check if run exists in database
            rr = client.table('run_results').select('run_id').eq('run_id', run_id).limit(1).execute()
            if not rr.data:
                raise HTTPException(status_code=404, detail='run not found')
        except Exception:
            raise HTTPException(status_code=404, detail='run not found')
    elif not run_exists_locally:
        raise HTTPException(status_code=404, detail='run not found')

    # TEA: DB only (no local fallback)
    tea: Dict[str, Any] | None = None
    try:
        if use_supabase_db():
            try:
                client = get_supabase()
                # Map run id
                db_run_id = run_id
                try:
                    r0 = client.table('runs').select('id,run_dir').eq('id', run_id).limit(1).execute()
                    if r0.data:
                        db_run_id = r0.data[0]['id']
                    else:
                        abs_path = str(RUNS / run_id)
                        r1 = client.table('runs').select('id,run_dir').eq('run_dir', abs_path).limit(1).execute()
                        if r1.data:
                            db_run_id = r1.data[0]['id']
                except Exception as e:
                    print(f"[WARN] Failed to resolve db_run_id for run_id={run_id}: {e}")
                    db_run_id = run_id
                r = client.table('run_persona_teas').select('*').eq('run_id', db_run_id).eq('persona_id', persona_id).limit(1).execute()
                print(f"TEA query result for {run_id}/{persona_id}: {r.data}")
                row = (r.data or [None])[0]
                if row:
                    tea = {
                        'thoughts': row.get('thoughts') or {},
                        'emotions': row.get('emotions') or {},
                        'hesitations': row.get('hesitations') or {},
                        'actions': row.get('actions') or {},
                        'sentiment_start': row.get('sentiment_start'),
                        'sentiment_end': row.get('sentiment_end'),
                    }
                    print(f"TEA data loaded: {tea}")
            except Exception as e:
                import traceback as _tb
                print(f"[ERROR] TEA query error for run_id={run_id} persona_id={persona_id}: {e}")
                _tb.print_exc()
                tea = None
    except Exception as e:
        import traceback as _tb
        print(f"[ERROR] TEA general error for run_id={run_id} persona_id={persona_id}: {e}")
        _tb.print_exc()
        tea = None

    # TEA fallback via llm_run_insights.persona_teas
    if tea is None and use_supabase_db():
        try:
            client = get_supabase()
            r = client.table('llm_run_insights').select('persona_teas').eq('run_id', run_id).limit(1).execute()
            print(f"LLM insights query result for {run_id}: {r.data}")
            row = (r.data or [None])[0]
            if row and isinstance(row.get('persona_teas'), dict):
                tea = row['persona_teas'].get(persona_id) or None
                print(f"LLM insights TEA data for persona {persona_id}: {tea}")
        except Exception as e:
            import traceback as _tb
            print(f"[ERROR] LLM insights query error for run_id={run_id}: {e}")
            _tb.print_exc()
            tea = None

    # Path distribution and exits/backtracks from local artifacts
    paths_count: Dict[str, int] = {}
    exits_count: Dict[str, int] = {}
    # Track backtracks by screen id primarily; keep name-only fallbacks when id is unavailable
    backtracks_by_screen_id: Dict[str, int] = {}
    backtracks_by_screen_name_fallback: Dict[str, int] = {}
    unique_thoughts: Dict[str, int] = {}
    backtracks_total = 0
    
    # Try database first for database-only runs
    if not run_exists_locally and use_supabase_db():
        try:
            client = get_supabase()

            # Attempt to load assets directory to map screen ids -> names
            screen_mapping: Dict[str, str] = {}
            try:
                rpath = client.table('runs').select('run_dir').eq('id', run_id).limit(1).execute()
                run_dir_str = (rpath.data or [{}])[0].get('run_dir') or ''
                if run_dir_str:
                    import pathlib as _pathlib, json as _json
                    nodes_path = _pathlib.Path(run_dir_str) / 'preprocess' / 'screen_nodes.json'
                    print(f"1Nodes path: {nodes_path}")
                    if nodes_path.exists():
                        nodes = _json.loads(nodes_path.read_text(encoding='utf-8'))
                        for node in nodes:
                            print(f"1Node: {node}")
                            sid = str(node.get('id', ''))
                            print(f"1Sid: {sid}")
                            if sid:
                                screen_mapping[sid] = str(node.get('name') or f"Screen {sid}")
            except Exception as e:
                print(f"1Error: {e}")
                # Non-fatal: fall back to id-based labels
                screen_mapping = {}

            # Get run_results for this persona
            rr = client.table('run_results').select('*').eq('run_id', run_id).eq('persona_id', persona_id).execute()
            rows = rr.data or []

            for r in rows:
                # Aggregate backtracks and exits
                backtracks_total += int(r.get('backtracks') or 0)
                status = (r.get('status') or '').lower()
                if status not in ['completed', '']:
                    dropoff_reason = r.get('dropoff_reason')
                    if dropoff_reason and str(dropoff_reason).lower() != 'completed':
                        exits_count[str(dropoff_reason)] = exits_count.get(str(dropoff_reason), 0) + 1

                # Build path from actions when available
                try:
                    actions_val = r.get('actions')
                    # actions may be already a list, or a JSON string
                    if isinstance(actions_val, str):
                        import json as _json
                        try:
                            actions = _json.loads(actions_val)
                        except Exception:
                            actions = []
                    else:
                        actions = actions_val or []

                    path_screens: list[str] = []
                    source_id = str(r.get('source_id') or '')
                    if source_id:
                        path_screens.append(screen_mapping.get(source_id, f"Screen {source_id}"))
                    if isinstance(actions, list) and actions:
                        for a in actions:
                            try:
                                to_id = str((a or {}).get('to_id') or '')
                                if to_id:
                                    name = screen_mapping.get(to_id, f"Screen {to_id}")
                                    if not path_screens or path_screens[-1] != name:
                                        path_screens.append(name)
                            except Exception:
                                continue
                    if path_screens:
                        label = ' > '.join(path_screens[:20])
                        paths_count[label] = paths_count.get(label, 0) + 1
                except Exception:
                    # Ignore malformed actions
                    pass

                # Extract thoughts
                thoughts_json = r.get('thoughts')
                if thoughts_json:
                    try:
                        thoughts_list = json.loads(thoughts_json) if isinstance(thoughts_json, str) else (thoughts_json or [])
                        for thought in thoughts_list:
                            if isinstance(thought, str) and thought.strip():
                                unique_thoughts[thought.strip()] = unique_thoughts.get(thought.strip(), 0) + 1
                    except Exception:
                        pass

            # Get friction/backtracks per screen
            try:
                fp = client.table('friction_points').select('*').eq('run_id', run_id).eq('persona_id', persona_id).execute()
                friction_rows = fp.data or []
                for fr in friction_rows:
                    sid = str(fr.get('screen_id') or '')
                    if sid:
                        backtracks_by_screen_id[sid] = backtracks_by_screen_id.get(sid, 0) + 1
            except Exception:
                pass

        except Exception as e:
            import traceback as _tb
            print(f"[ERROR] Loading database data for persona detail failed run_id={run_id} persona_id={persona_id}: {e}")
            _tb.print_exc()
    
    # Fallback to local files if they exist (support DB-resolved run_dir too)
    try:
        run_dir = RUNS / run_id
        if not run_dir.exists() and use_supabase_db():
            try:
                client = get_supabase()
                # Try exact id -> run_dir
                r0 = client.table('runs').select('run_dir').eq('id', run_id).limit(1).execute()
                cand = (r0.data or [{}])[0].get('run_dir')
                if not cand:
                    # Fallback: find a run_dir that ends with this run_id (some deployments store absolute run_dir path)
                    r1 = client.table('runs').select('run_dir').like('run_dir', f"%/{run_id}").limit(1).execute()
                    cand = (r1.data or [{}])[0].get('run_dir')
                if cand:
                    import pathlib as _pathlib
                    cpath = _pathlib.Path(cand)
                    if cpath.exists():
                        run_dir = cpath
            except Exception:
                pass
        # persona summary for quick stats
        psum = run_dir / 'tests' / 'persona_summary.json'
        if psum.exists():
            try:
                import json as _json
                data = _json.loads(psum.read_text(encoding='utf-8'))
                for r in (data.get('results') or []):
                    if str(r.get('persona_id')) != str(persona_id):
                        continue
                    backtracks_total += int(r.get('backtracks') or 0)
                    # Dropoff reason - only for sessions that did NOT complete
                    status = str(r.get('status') or '').lower()
                    if status not in ['completed', '']:
                        dps = list(r.get('drop_off_points') or [])
                        if dps:
                            reason = str((dps[-1] or {}).get('reason') or r.get('status') or 'unknown')
                            if reason.lower() != 'completed':
                                exits_count[reason] = exits_count.get(reason, 0) + 1
            except Exception:
                pass
        # Load screen mapping for ID to name conversion
        screen_mapping = {}
        try:
            screen_nodes_path = run_dir / 'preprocess' / 'screen_nodes.json'
            print(f"2Screen nodes path: {screen_nodes_path}")
            if screen_nodes_path.exists():
                import json as _json
                screen_nodes = _json.loads(screen_nodes_path.read_text(encoding='utf-8'))
                for node in screen_nodes:
                    print(f"2Node: {node}")
                    screen_mapping[str(node.get('id', ''))] = str(node.get('name', f"Screen {node.get('id', '')}"))
        except Exception as e:
            print(f"2Error: {e}")
            pass

        # path.json + user_report.json per simulation
        sim_root = run_dir / 'tests'
        for pdir in (sim_root.glob(f'persona_{persona_id}/simulations/*')):
            try:
                import json as _json
                # Build a single path label per simulation: prefer user_report.json actions; fallback to path.json screens
                label_for_sim = None
                try:
                    ureport = pdir / 'user_report.json'
                    if ureport.exists():
                        ur = _json.loads(ureport.read_text(encoding='utf-8'))
                        actions = ur.get('actions') or []
                        if isinstance(actions, list) and actions:
                            path_screens = []
                            source_id = str(ur.get('source_id', ''))
                            # Only attempt mapping if available in local scope
                            if source_id and ('screen_mapping' in locals()) and (source_id in screen_mapping):
                                path_screens.append(screen_mapping[source_id])
                            # Append destination names
                            for action in actions:
                                to_id = str(action.get('to_id', ''))
                                if not to_id:
                                    continue
                                if ('screen_mapping' in locals()) and (to_id in screen_mapping):
                                    screen_name = screen_mapping[to_id]
                                else:
                                    screen_name = f"Screen {to_id}"
                                if not path_screens or path_screens[-1] != screen_name:
                                    path_screens.append(screen_name)
                            if path_screens:
                                label_for_sim = ' > '.join(path_screens[:20])
                    if not label_for_sim:
                        ppath = pdir / 'path.json'
                        if ppath.exists():
                            pobj = _json.loads(ppath.read_text(encoding='utf-8'))
                            screens = pobj.get('screens') or []
                            if screens:
                                label_for_sim = ' > '.join(str(s) for s in screens[:20])
                    if label_for_sim:
                        paths_count[label_for_sim] = paths_count.get(label_for_sim, 0) + 1
                except Exception:
                    pass

                # Also collect backtracks/thoughts from user_report.json when present
                try:
                    if 'ur' not in locals():
                        ur = None
                        ureport = pdir / 'user_report.json'
                        if ureport.exists():
                            ur = _json.loads(ureport.read_text(encoding='utf-8'))
                    if isinstance(ur, dict):
                        for fp in ur.get('friction_points') or []:
                            if fp.get('type') == 'back_or_close':
                                sid = str(fp.get('screen_id') or '')
                                if sid:
                                    backtracks_by_screen_id[sid] = backtracks_by_screen_id.get(sid, 0) + 1
                                else:
                                    nm = str(fp.get('screen_name') or 'Unknown')
                                    backtracks_by_screen_name_fallback[nm] = backtracks_by_screen_name_fallback.get(nm, 0) + 1
                        for b in ur.get('backtracks') or []:
                            sid = str(b.get('screen_id') or '')
                            if sid:
                                backtracks_by_screen_id[sid] = backtracks_by_screen_id.get(sid, 0) + 1
                            else:
                                nm = str(b.get('screen_name') or 'Unknown')
                                backtracks_by_screen_name_fallback[nm] = backtracks_by_screen_name_fallback.get(nm, 0) + 1
                        for t in ur.get('thoughts') or []:
                            s = (str(t or '').strip())
                            if s:
                                unique_thoughts[s] = unique_thoughts.get(s, 0) + 1
                except Exception:
                    pass
            except Exception:
                continue
    except Exception as e:
        import traceback as _tb
        print(f"[ERROR] Persona detail local files scan failed run_id={run_id} persona_id={persona_id}: {e}")
        _tb.print_exc()

    # Normalize path distribution (top 5)
    total_paths = sum(paths_count.values()) or 1
    paths = [
        {'path': k, 'count': v, 'sharePct': round((100.0 * v) / total_paths, 1)}
        for k, v in sorted(paths_count.items(), key=lambda kv: kv[1], reverse=True)[:5]
    ]
    exits = [
        {'reason': k, 'count': v}
        for k, v in sorted(exits_count.items(), key=lambda kv: kv[1], reverse=True)
    ]
    
    # If no drop-off reasons found, provide a helpful message
    if not exits:
        exits = [{'reason': 'No drop-offs recorded', 'count': 0}]
    # Build backtracks list with both screen_id and screen name
    # First, enrich id-based counts with names when available
    backs: List[Dict[str, Any]] = []
    try:
        # Build id->name map from any loaded mapping (id_to_name comes later), so prepare a temporary map
        id_to_name_temp: Dict[str, str] = {}
        try:
            # Prefer local mapping if we have already loaded nodes
            for nid_int, nm in (id_to_name.items() if 'id_to_name' in locals() else []):
                id_to_name_temp[str(nid_int)] = nm
        except Exception:
            pass
        # Merge id-based counts
        for sid, cnt in sorted(backtracks_by_screen_id.items(), key=lambda kv: kv[1], reverse=True):
            backs.append({
                'id': (int(sid) if str(sid).isdigit() else None),
                'screen_id': sid,
                'screen': id_to_name_temp.get(sid, sid),
                'count': cnt,
            })
        # Add name-only fallbacks (no screen_id available)
        for nm, cnt in sorted(backtracks_by_screen_name_fallback.items(), key=lambda kv: kv[1], reverse=True):
            backs.append({'id': None, 'screen_id': None, 'screen': nm, 'count': cnt})
    except Exception:
        # As a final fallback, collapse to name-only map if something goes wrong
        tmp = {}
        for k, v in backtracks_by_screen_id.items():
            tmp[k] = tmp.get(k, 0) + v
        for k, v in backtracks_by_screen_name_fallback.items():
            tmp[k] = tmp.get(k, 0) + v
        backs = [ {'id': (int(k) if k.isdigit() else None), 'screen_id': (k if k.isdigit() else None), 'screen': k, 'count': v} for k, v in sorted(tmp.items(), key=lambda kv: kv[1], reverse=True) ]
    # Map numeric screen ids to friendly names when local preprocess artifacts exist
    screen_files: List[Dict[str, Any]] = []
    id_to_name: Dict[int, str] = {}
    id_to_file: Dict[int, str] = {}
    
    try:
        # Resolve run_dir for DB-only runs using Supabase metadata
        try:
            import pathlib as _pathlib
            if not (RUNS / run_id).exists() and use_supabase_db():
                try:
                    client = get_supabase()
                    r0 = client.table('runs').select('run_dir,id').eq('id', run_id).limit(1).execute()
                    _rd = None
                    if r0.data and r0.data[0].get('run_dir'):
                        _rd = r0.data[0]['run_dir']
                    else:
                        r1 = client.table('runs').select('run_dir,id').like('run_dir', f"%/{run_id}").limit(1).execute()
                        if r1.data and r1.data[0].get('run_dir'):
                            _rd = r1.data[0]['run_dir']
                    if _rd:
                        try:
                            run_dir = _pathlib.Path(_rd)
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception as e:
            import traceback as _tb
            print(f"[WARN] Failed to resolve assets base dir from DB for run_id={run_id}: {e}")
            _tb.print_exc()

        # Try to load screen mapping from local files (works for both local and database-only runs)
        # Prefer the resolved run_dir from the DB when available; otherwise fall back to RUNS/<run_id>
        assets_base_dir = None
        try:
            import pathlib as _pathlib
            rd_candidate = run_dir  # may have been resolved from DB above
        except Exception:
            rd_candidate = None
        try:
            for cand in [rd_candidate, RUNS / run_id]:
                if cand and _pathlib.Path(cand).exists():
                    assets_base_dir = _pathlib.Path(cand)
                    break
        except Exception:
            assets_base_dir = RUNS / run_id

        nodes_path = assets_base_dir / 'preprocess' / 'screen_nodes.json'
        print(f"3Nodes path: {nodes_path}")
        if nodes_path.exists():
            try:
                nodes = json.loads(nodes_path.read_text(encoding='utf-8'))
                for n in nodes or []:
                    try:
                        print(f"3Node: {n}")
                        nid = int(n.get('id'))
                        nm = str(n.get('name') or '')
                        fn = str(n.get('file') or '')
                        print(f"3Nid: {nid}")
                        print(f"3Nm: {nm}")
                        print(f"3Fn: {fn}")
                        if isinstance(nid, int) and nm:
                            id_to_name[nid] = nm
                        if isinstance(nid, int) and fn:
                            id_to_file[nid] = fn
                    except Exception:
                        continue
            except Exception as e:
                import traceback as _tb
                print(f"[WARN] Failed to load screen_nodes.json for run_id={run_id} from {nodes_path}: {e}")
                _tb.print_exc()
        
        # Rebuild backs now that we have id_to_name mapping to ensure ids are filled when possible
        try:
            def _norm(s: str) -> str:
                return " ".join(str(s or '').split()).strip().lower()
            # Build name->id map (pick smallest id on duplicate names)
            name_to_id: Dict[str, int] = {}
            for nid, nm in id_to_name.items():
                key = _norm(nm)
                if key in name_to_id:
                    name_to_id[key] = min(name_to_id[key], int(nid))
                else:
                    name_to_id[key] = int(nid)

            backs_enriched: List[Dict[str, Any]] = []
            for sid, cnt in sorted(backtracks_by_screen_id.items(), key=lambda kv: kv[1], reverse=True):
                try:
                    nid = int(sid) if str(sid).isdigit() else None
                except Exception:
                    nid = None
                label = id_to_name.get(nid) if (nid is not None and nid in id_to_name) else (sid if sid else None)
                backs_enriched.append({'id': nid, 'screen_id': (str(nid) if nid is not None else None), 'screen': (label or ''), 'count': cnt})

            for nm, cnt in sorted(backtracks_by_screen_name_fallback.items(), key=lambda kv: kv[1], reverse=True):
                key = _norm(nm)
                nid = name_to_id.get(key)
                backs_enriched.append({'id': (nid if nid is not None else None), 'screen_id': (str(nid) if nid is not None else None), 'screen': nm, 'count': cnt})

            backs = backs_enriched
        except Exception:
            pass
        
        # Build screen files list for all screens (not just those with backtracks)
        for nid, fn in id_to_file.items():
            # Build URL relative to the chosen assets base dir
            p = (assets_base_dir / 'preprocess' / 'screens' / fn)
            if p.exists():
                public_url: str | None = None
                try:
                    # Try to upload and get a signed public URL when using Supabase
                    from ..storage import upload_file_to_supabase
                    from ..metrics import slugify as _slug
                    proj_slug = _slug(project_name if 'project_name' in locals() else 'project')
                    storage_path = f"{proj_slug}/runs/{run_id}/preprocess/screens/{fn}"
                    public_url = upload_file_to_supabase(p, storage_path)
                except Exception:
                    public_url = None
                screen_files.append({
                    'id': nid,
                    'name': id_to_name.get(nid) or str(nid),
                    # Prefer signed public URL; else expose via /runs-files when path is under RUNS
                    'image': (public_url or ((f"/runs-files/{p.relative_to(RUNS)}") if str(p).startswith(str(RUNS)) else None))
                })
            else:
                # File doesn't exist, add entry without image
                screen_files.append({
                    'id': nid,
                    'name': id_to_name.get(nid) or str(nid),
                    'image': None
                })
    except Exception as e:
        import traceback as _tb
        print(f"[WARN] Screen file processing failed for run_id={run_id}: {e}")
        _tb.print_exc()
    
    # Final pass: if any backs entries are missing id/screen_id, try resolving via screen_files names
    try:
        if screen_files:
            # Build map from normalized name -> sorted list of ids (prefer smallest)
            def _norm2(s: str) -> str:
                return " ".join(str(s or '').split()).strip().lower()
            name_to_ids: Dict[str, List[int]] = {}
            for sf in screen_files:
                try:
                    nid = int(sf.get('id'))
                    nm = str(sf.get('name') or '')
                    key = _norm2(nm)
                    if key not in name_to_ids:
                        name_to_ids[key] = []
                    name_to_ids[key].append(nid)
                except Exception:
                    continue
            for k in name_to_ids:
                name_to_ids[k] = sorted(set(name_to_ids[k]))
            for b in backs:
                if b.get('id') is None:
                    nm = b.get('screen')
                    if isinstance(nm, str):
                        key = _norm2(nm)
                        ids = name_to_ids.get(key) or []
                        if ids:
                            nid = ids[0]
                            b['id'] = nid
                            b['screen_id'] = str(nid)
    except Exception:
        pass

    thoughts = [
        {'text': k, 'count': v}
        for k, v in sorted(unique_thoughts.items(), key=lambda kv: kv[1], reverse=True)[:20]
    ]

    # Build Unique TEA Thoughts timeline for this persona from latest simulation's journey.jsonl
    tea_thoughts: List[Dict[str, Any]] = []
    latest_sim = None  # Declare outside try block for reuse
    figma_to_id: Dict[str, int] = {}
    
    try:
        # Build figma screen id -> local numeric id map using already loaded nodes
        try:
            nodes_path = run_dir / 'preprocess' / 'screen_nodes.json'
            if nodes_path.exists():
                nodes = json.loads(nodes_path.read_text(encoding='utf-8'))
                for n in (nodes or []):
                    try:
                        fid = str(n.get('screen_id') or '')
                        lid = int(n.get('id')) if str(n.get('id') or '').isdigit() else None
                        if fid and isinstance(lid, int):
                            figma_to_id[fid] = lid
                    except Exception:
                        continue
        except Exception:
            figma_to_id = {}

        # Find latest simulation folder for this persona
        sim_root = run_dir / 'tests' / f'persona_{persona_id}' / 'simulations'
        print(f"[DEBUG TEA] run_dir={run_dir}, sim_root={sim_root}, exists={sim_root.exists()}")
        if sim_root.exists():
            try:
                sims = [p for p in sim_root.iterdir() if p.is_dir()]
                print(f"[DEBUG TEA] Found {len(sims)} simulations")
                if sims:
                    latest_sim = sorted(sims)[-1]
                    print(f"[DEBUG TEA] Latest sim: {latest_sim}")
            except Exception as e:
                print(f"[DEBUG TEA] Error finding sims: {e}")
                latest_sim = None
        else:
            print(f"[DEBUG TEA] sim_root does not exist")
        # Parse journey.jsonl for step-by-step thoughts
        if latest_sim:
            jpath = latest_sim / 'journey.jsonl'
            if jpath.exists():
                try:
                    # Helper: light emotion palette
                    def _emo_style(name: str) -> Dict[str, Any]:
                        n = (name or '').strip().lower()
                        palette = [
                            ('excited', '💥', '#EF4444'),
                            ('joy', '😊', '#F59E0B'),
                            ('curious', '🧐', '#F59E0B'),
                            ('cautious', '🤔', '#FCD34D'),
                            ('focused', '🎯', '#34D399'),
                            ('confused', '😕', '#F97316'),
                            ('frustrated', '😣', '#F87171'),
                            ('neutral', '😐', '#94A3B8'),
                        ]
                        for key, emoji, color in palette:
                            if key in n:
                                return {'name': name, 'emoji': emoji, 'color': color}
                        return {'name': name or 'Neutral', 'emoji': '😐', 'color': '#94A3B8'}

                    lines = jpath.read_text(encoding='utf-8').splitlines()
                    for idx, line in enumerate(lines):
                        if len(tea_thoughts) >= 30:
                            break
                        try:
                            step = json.loads(line)
                        except Exception:
                            continue
                        # Map figma screen id string -> local numeric id for assets
                        sid_str = str(step.get('screen_id') or '')
                        lid = None
                        try:
                            # Some runs may already store numeric id as string
                            lid = int(sid_str) if sid_str.isdigit() else figma_to_id.get(sid_str)
                        except Exception:
                            lid = figma_to_id.get(sid_str)
                        # Resolve names and thumbnail
                        screen_name = id_to_name.get(int(lid)) if isinstance(lid, int) else (step.get('frame_name') or None)
                        thumb = None
                        if isinstance(lid, int) and lid in id_to_file:
                            img_rel = (run_dir / 'preprocess' / 'screens' / id_to_file[lid])
                            if img_rel.exists():
                                try:
                                    thumb = f"/runs-files/{img_rel.relative_to(RUNS)}"
                                except Exception:
                                    thumb = None
                        # Choose thought text
                        thought = ''
                        mt = step.get('micro_thoughts') or []
                        if isinstance(mt, list) and mt:
                            thought = str(mt[0])
                        if not thought:
                            thought = str(step.get('goal_based_narrative') or step.get('narrative') or '').strip()
                        # Build emotion badge
                        emo_name = str(step.get('emotion') or step.get('second_emotion') or '').strip()
                        emotion = _emo_style(emo_name)
                        # Traits not available here; keep optional
                        entry = {
                            'screen_id': (int(lid) if isinstance(lid, int) else None),
                            'screen_name': screen_name or (sid_str or f'Step {idx+1}'),
                            'screen_thumbnail_url': thumb,
                            'thought_text': thought,
                            'emotion': emotion,
                            'goal': None,
                            'action': (step.get('final_action') or step.get('first_action') or ''),
                            'traits': None,
                            'friction': float(step.get('severity_weighted_friction') or 0.0),
                            'success': bool(step.get('is_goal_screen')),
                        }
                        tea_thoughts.append(entry)
                except Exception:
                    tea_thoughts = []
    except Exception:
        tea_thoughts = []

    # Build Emotion Mix data (aggregated counts + journey) from journey.jsonl
    emotions: List[Dict[str, Any]] = []
    emotion_journey: List[Dict[str, Any]] = []
    sentiment_series: List[Dict[str, Any]] = []
    sentiment_events: List[Dict[str, Any]] = []
    try:
        # Reuse latest_sim from tea_thoughts logic above
        if latest_sim:
            jpath = latest_sim / 'journey.jsonl'
            if jpath.exists():
                try:
                    # Helper: emotion style with valence
                    def _emo_style_full(name: str) -> Dict[str, Any]:
                        n = (name or '').strip().lower()
                        palette = [
                            ('excited', '💥', '#EF4444', 0.9),
                            ('joy', '😊', '#F59E0B', 0.95),
                            ('curious', '🧐', '#FBD35D', 0.8),
                            ('cautious', '🤔', '#FCD34D', 0.6),
                            ('focused', '🎯', '#34D399', 0.85),
                            ('calm', '🧘', '#6EE7B7', 0.7),
                            ('confident', '😎', '#3B82F6', 0.9),
                            ('confused', '😕', '#F97316', 0.3),
                            ('frustrated', '😣', '#F87171', 0.2),
                            ('neutral', '😐', '#94A3B8', 0.5),
                        ]
                        for key, emoji, color, valence in palette:
                            if key in n:
                                return {'name': name, 'emoji': emoji, 'color': color, 'valence': valence}
                        return {'name': name or 'Neutral', 'emoji': '😐', 'color': '#94A3B8', 'valence': 0.5}
                    
                    emotion_counts: Dict[str, int] = {}
                    emotion_meta: Dict[str, Dict[str, Any]] = {}
                    
                    lines = jpath.read_text(encoding='utf-8').splitlines()
                    for idx, line in enumerate(lines):
                        try:
                            step = json.loads(line)
                        except Exception:
                            continue
                        emo_name = str(step.get('emotion') or step.get('second_emotion') or '').strip()
                        if not emo_name:
                            emo_name = 'Neutral'
                        # Aggregate counts
                        emotion_counts[emo_name] = emotion_counts.get(emo_name, 0) + 1
                        if emo_name not in emotion_meta:
                            emotion_meta[emo_name] = _emo_style_full(emo_name)
                        # Build journey step
                        sid_str = str(step.get('screen_id') or '')
                        lid = None
                        try:
                            lid = int(sid_str) if sid_str.isdigit() else figma_to_id.get(sid_str)
                        except Exception:
                            lid = figma_to_id.get(sid_str)
                        screen_name = id_to_name.get(int(lid)) if isinstance(lid, int) else (step.get('frame_name') or 'Screen')
                        emo_style = _emo_style_full(emo_name)
                        emotion_journey.append({
                            'screen': screen_name,
                            'emotion': emo_name,
                            'color': emo_style['color'],
                            'emoji': emo_style['emoji'],
                        })
                        # Sentiment series entry
                        try:
                            sentiment_series.append({
                                'idx': idx,
                                'valence': float(emo_style.get('valence') or 0.5),
                                'screen_id': (int(lid) if isinstance(lid, int) else None),
                                'screen_name': screen_name,
                            })
                        except Exception:
                            pass
                        # Simple event markers
                        try:
                            if bool(step.get('backtrack_flag')):
                                sentiment_events.append({'idx': idx, 'type': 'backtrack', 'severity': 3})
                        except Exception:
                            pass
                    
                    # Build emotions summary
                    for emo_name, count in sorted(emotion_counts.items(), key=lambda kv: kv[1], reverse=True):
                        meta = emotion_meta.get(emo_name, _emo_style_full(emo_name))
                        emotions.append({
                            'name': emo_name,
                            'count': count,
                            'color': meta['color'],
                            'emoji': meta['emoji'],
                            'valence': meta['valence'],
                        })
                except Exception:
                    emotions = []
                    emotion_journey = []
                    sentiment_series = []
                    sentiment_events = []
    except Exception:
        emotions = []
        emotion_journey = []
        sentiment_series = []
        sentiment_events = []

    # Derive sentiment start/end if missing
    try:
        if (tea is None) or (tea.get('sentiment_start') is None) or (tea.get('sentiment_end') is None):
            if sentiment_series:
                s0 = float(sentiment_series[0].get('valence') or 0.0)
                s1 = float(sentiment_series[-1].get('valence') or 0.0)
                tea = (tea or {})
                tea.setdefault('sentiment_start', s0)
                tea.setdefault('sentiment_end', s1)
    except Exception:
        pass

    # Analyze Unique TEAs (outliers) for edge case detection
    unique_teas = await _analyze_unique_teas(run_id, persona_id)

    return {
        'run_id': run_id,
        'persona_id': persona_id,
        'tea': tea or {},
        'paths': paths,
        'backtracks': backtracks_total,
        'exits': exits,
        'backtracks_by_screen': backs,
        'screen_files': screen_files,
        'thoughts_unique': thoughts,
        'unique_teas': unique_teas,
        'tea_thoughts': tea_thoughts,
        'emotions': emotions,
        'emotion_journey': emotion_journey,
        'sentiment_series': sentiment_series,
        'sentiment_events': sentiment_events,
    }
# _severity_for_category now imported from .utils


@router.get('/runs/{run_id}/persona/{persona_id}/users.csv')
async def persona_users_csv(run_id: str, persona_id: str, format: str = 'csv'):
    """Download the concrete user list used for a persona slot in a run.

    Args:
        run_id: The run identifier
        persona_id: The persona slot identifier  
        format: Export format - 'csv' for enhanced CSV or 'xlsx' for Excel
    
    Output formats:
        CSV: Enhanced CSV with persona summary section and user details
        XLSX: Multi-sheet Excel with Persona Info, Users, and Statistics sheets
    """
    # Determine run_dir and get project name
    run_dir = RUNS / run_id
    project_name = 'Unknown'
    
    if not run_dir.exists() and use_supabase_db():
        try:
            client = get_supabase()
            r0 = client.table('runs').select('run_dir,id').eq('id', run_id).limit(1).execute()
            if r0.data and r0.data[0].get('run_dir'):
                run_dir = pathlib.Path(r0.data[0]['run_dir'])
            else:
                r1 = client.table('runs').select('run_dir,id').like('run_dir', f"%/{run_id}").limit(1).execute()
                if r1.data and r1.data[0].get('run_dir'):
                    run_dir = pathlib.Path(r1.data[0]['run_dir'])
        except Exception:
            pass
    
    # Try to get project name
    try:
        from .status_full import read_projects
        projects = read_projects()
        for project in projects:
            if project.get('run_id') == run_id or project.get('id') == run_id:
                project_name = project.get('name', 'Unknown')
                break
    except Exception:
        pass
        
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail='run_dir not found')

    # Load resolution file
    res_path = run_dir / 'tests' / 'persona_resolution.json'
    if not res_path.exists():
        raise HTTPException(status_code=404, detail='persona resolution not found')
    try:
        res = json.loads(res_path.read_text(encoding='utf-8'))
    except Exception:
        raise HTTPException(status_code=500, detail='failed reading persona resolution')
    sel = None
    for p in (res.get('personas') or []):
        if str(p.get('slot')) == str(persona_id):
            sel = p
            break
    if not sel:
        raise HTTPException(status_code=404, detail='persona slot not found')

    # Load users master
    users_path = ROOT / 'users' / 'users.json'
    if not users_path.exists():
        for cand in [pathlib.Path.cwd() / 'users' / 'users.json', pathlib.Path('/Users/ankita/Documents/workspace/design-agent-simulator/users/users.json')]:
            if cand.exists():
                users_path = cand
                break
    if not users_path.exists():
        raise HTTPException(status_code=500, detail='persona.json not found')
    try:
        users = json.loads(users_path.read_text(encoding='utf-8'))
    except Exception:
        raise HTTPException(status_code=500, detail='failed parsing persona.json')
    by_id = {str(u.get('id')): u for u in users if 'id' in u}

    def ocean(u: Dict[str, Any], k: str):
        try:
            v = (u.get('ocean') or {}).get(k)
            if isinstance(v, dict):
                v = v.get('value')
            return float(v)
        except Exception:
            return ''

    # Get persona statistics
    persona_stats = await _get_persona_statistics(run_dir, persona_id)
    
    pname = str(sel.get('name') or f"Persona {sel.get('slot')}")
    traits = str(sel.get('traits') or '')
    user_ids = sel.get('user_ids') or []
    
    # Generate filename
    safe_pname = "".join(c for c in pname if c.isalnum() or c in (' ', '-', '_')).rstrip()
    safe_project = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    
    if format.lower() == 'xlsx':
        return await _generate_excel_export(
            run_id, pname, traits, sel.get('slot'), user_ids, by_id, 
            persona_stats, safe_pname, safe_project
        )
    else:
        return await _generate_csv_export(
            run_id, pname, traits, sel.get('slot'), user_ids, by_id,
            persona_stats, safe_pname, safe_project
        )

async def _get_persona_statistics(run_dir: pathlib.Path, persona_id: str) -> Dict[str, Any]:
    """Get statistics for a persona from the run data."""
    stats = {
        'total_users': 0,
        'completed_users': 0,
        'completion_rate': 0.0,
        'average_steps': 0.0,
        'average_time_seconds': 0.0
    }
    
    try:
        # Try to load from persona_summary.csv
        summary_path = run_dir / 'tests' / 'persona_summary.csv'
        if summary_path.exists():
            import csv as csv_module
            with open(summary_path, 'r', encoding='utf-8') as f:
                reader = csv_module.DictReader(f)
                for row in reader:
                    if str(row.get('persona_id', '')) == str(persona_id):
                        stats['total_users'] += 1
                        if row.get('status', '').lower() == 'completed':
                            stats['completed_users'] += 1
                        
                        try:
                            steps = float(row.get('steps', 0))
                            time_sec = float(row.get('time_sec', 0))
                            stats['average_steps'] += steps
                            stats['average_time_seconds'] += time_sec
                        except (ValueError, TypeError):
                            pass
            
            if stats['total_users'] > 0:
                stats['completion_rate'] = (stats['completed_users'] / stats['total_users']) * 100
                stats['average_steps'] = stats['average_steps'] / stats['total_users']
                stats['average_time_seconds'] = stats['average_time_seconds'] / stats['total_users']
    except Exception:
        pass
    
    return stats

async def _generate_csv_export(
    run_id: str, pname: str, traits: str, slot: int, user_ids: list, 
    by_id: Dict[str, Any], persona_stats: Dict[str, Any], 
    safe_pname: str, safe_project: str
) -> Response:
    """Generate enhanced CSV export with persona summary section."""
    sio = io.StringIO()
    wr = csv.writer(sio)
    
    # Persona Information Section
    wr.writerow(['PERSONA INFORMATION'])
    wr.writerow(['Persona Name', pname])
    wr.writerow(['Persona Slot', slot])
    wr.writerow(['Traits', traits])
    wr.writerow(['Total Users', persona_stats['total_users']])
    wr.writerow(['Completion Rate', f"{persona_stats['completion_rate']:.1f}%"])
    wr.writerow(['Average Steps', f"{persona_stats['average_steps']:.1f}"])
    wr.writerow(['Average Time (seconds)', f"{persona_stats['average_time_seconds']:.1f}"])
    wr.writerow([])  # Empty row
    
    # User Details Section
    wr.writerow(['USER DETAILS'])
    # Removed 'User ID' as requested; keep remaining headers
    wr.writerow(['Name', 'Job', 'Country', 'Bio', 'Openness', 'Conscientiousness', 'Extraversion', 'Agreeableness', 'Neuroticism'])
    
    def ocean(u: Dict[str, Any], k: str):
        try:
            v = (u.get('ocean') or {}).get(k)
            if isinstance(v, dict):
                v = v.get('value')
            return float(v)
        except Exception:
            return ''
    
    for uid in user_ids:
        u = by_id.get(str(uid), {})
        wr.writerow([
            u.get('name',''), 
            u.get('job',''), 
            u.get('country',''),
            (u.get('bio') or u.get('description','')),
            ocean(u,'O'), 
            ocean(u,'C'), 
            ocean(u,'E'), 
            ocean(u,'A'), 
            ocean(u,'N')
        ])
    
    buf = sio.getvalue().encode('utf-8')
    filename = f"Users_{safe_pname}_{run_id}_{safe_project}.csv"
    
    return Response(content=buf, media_type='text/csv; charset=utf-8', headers={
        'Content-Disposition': f'attachment; filename="{filename}"',
        'Cache-Control': 'no-store',
    })

async def _generate_excel_export(
    run_id: str, pname: str, traits: str, slot: int, user_ids: list,
    by_id: Dict[str, Any], persona_stats: Dict[str, Any],
    safe_pname: str, safe_project: str
) -> Response:
    """Generate two-sheet Excel export with left-aligned content.

    Sheets:
      - Persona Summary (column-wise single-row summary)
      - User Details (rows for each user; no User ID column)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Sheet 1: Persona Summary
        ws_summary = wb.create_sheet("Persona Summary")
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        ws_summary.append(['Persona Name', 'Persona Slot', 'Traits', 'Total Users', 'Completion Rate (%)', 'Average Steps', 'Average Time (seconds)'])
        ws_summary.append([
            pname,
            slot,
            traits,
            persona_stats['total_users'],
            float(f"{persona_stats['completion_rate']:.1f}"),
            float(f"{persona_stats['average_steps']:.1f}"),
            float(f"{persona_stats['average_time_seconds']:.1f}")
        ])
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left', vertical='top')
        
        # Sheet 2: User Details
        ws_users = wb.create_sheet("User Details")
        headers = ['Name', 'Job', 'Country', 'Bio', 'Openness', 'Conscientiousness', 'Extraversion', 'Agreeableness', 'Neuroticism']
        ws_users.append(headers)
        
        # Format headers
        for cell in ws_users[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='left', vertical='top')
        
        def ocean(u: Dict[str, Any], k: str):
            try:
                v = (u.get('ocean') or {}).get(k)
                if isinstance(v, dict):
                    v = v.get('value')
                return float(v)
            except Exception:
                return ''
        
        for uid in user_ids:
            u = by_id.get(str(uid), {})
            ws_users.append([
                u.get('name',''), 
                u.get('job',''), 
                u.get('country',''),
                (u.get('bio') or u.get('description','')),
                ocean(u,'O'), 
                ocean(u,'C'), 
                ocean(u,'E'), 
                ocean(u,'A'), 
                ocean(u,'N')
            ])

        # Left align all cells
        for ws in (ws_summary, ws_users):
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Auto-adjust column widths for both sheets
        for sheet in (ws_summary, ws_users):
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"Users_{safe_pname}_{run_id}_{safe_project}.xlsx"
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Cache-Control': 'no-store',
            }
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail='Excel export requires openpyxl library')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Excel export failed: {str(e)}')

@router.get('/runs/{run_id}/users.xlsx')
async def all_personas_users_excel(run_id: str):
    """Download all personas' user data in a multi-sheet Excel file.

    Creates an Excel file with:
    - Sheet 1: "Persona Summary" - Overview of all personas
    - Sheet 2: "All Users" - All users from all personas
    """
    # Determine run_dir and get project name
    run_dir = RUNS / run_id
    project_name = 'Unknown'
    
    if not run_dir.exists() and use_supabase_db():
        try:
            client = get_supabase()
            r0 = client.table('runs').select('run_dir,id').eq('id', run_id).limit(1).execute()
            if r0.data and r0.data[0].get('run_dir'):
                run_dir = pathlib.Path(r0.data[0]['run_dir'])
            else:
                r1 = client.table('runs').select('run_dir,id').like('run_dir', f"%/{run_id}").limit(1).execute()
                if r1.data and r1.data[0].get('run_dir'):
                    run_dir = pathlib.Path(r1.data[0]['run_dir'])
        except Exception:
            pass
    
    # Try to get project name
    try:
        from .status_full import read_projects
        projects = read_projects()
        for project in projects:
            if project.get('run_id') == run_id or project.get('id') == run_id:
                project_name = project.get('name', 'Unknown')
                break
    except Exception:
        pass
        
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail='run_dir not found')

    # Load resolution file
    res_path = run_dir / 'tests' / 'persona_resolution.json'
    if not res_path.exists():
        raise HTTPException(status_code=404, detail='persona resolution not found')
    
    try:
        res = json.loads(res_path.read_text(encoding='utf-8'))
    except Exception:
        raise HTTPException(status_code=500, detail='failed reading persona resolution')

    # Load users master
    users_path = ROOT / 'users' / 'users.json'
    if not users_path.exists():
        for cand in [pathlib.Path.cwd() / 'users' / 'users.json', pathlib.Path('/Users/ankita/Documents/workspace/design-agent-simulator/users/users.json')]:
            if cand.exists():
                users_path = cand
                break
    if not users_path.exists():
        raise HTTPException(status_code=500, detail='users.json not found')
    
    try:
        users = json.loads(users_path.read_text(encoding='utf-8'))
    except Exception:
        raise HTTPException(status_code=500, detail='failed parsing users.json')
    
    by_id = {str(u.get('id')): u for u in users if 'id' in u}
    
    def ocean(u: Dict[str, Any], k: str):
        try:
            v = (u.get('ocean') or {}).get(k)
            if isinstance(v, dict):
                v = v.get('value')
            return float(v)
        except Exception:
            return ''

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Persona Summary
        ws_summary = wb.create_sheet("Persona Summary")
        ws_summary.append(['Persona Name', 'Persona Slot', 'Traits', 'Total Users', 'Completion Rate', 'Average Steps', 'Average Time (seconds)'])
        
        # Sheet 2: All Users
        ws_users = wb.create_sheet("All Users")
        ws_users.append(['Persona Name', 'Persona Slot', 'User ID', 'Name', 'Job', 'Country', 'Bio', 'Openness', 'Conscientiousness', 'Extraversion', 'Agreeableness', 'Neuroticism'])
        
        # Format headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for ws in [ws_summary, ws_users]:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
        
        personas = res.get('personas', [])
        for persona in personas:
            persona_id = str(persona.get('slot', ''))
            pname = str(persona.get('name') or f"Persona {persona.get('slot')}")
            traits = str(persona.get('traits') or '')
            user_ids = persona.get('user_ids', [])
            
            # Get statistics for this persona
            persona_stats = await _get_persona_statistics(run_dir, persona_id)
            
            # Add to summary sheet
            ws_summary.append([
                pname,
                persona.get('slot'),
                traits,
                persona_stats['total_users'],
                f"{persona_stats['completion_rate']:.1f}%",
                f"{persona_stats['average_steps']:.1f}",
                f"{persona_stats['average_time_seconds']:.1f}"
            ])
            
            # Add users to users sheet
            for uid in user_ids:
                u = by_id.get(str(uid), {})
                ws_users.append([
                    pname,
                    persona.get('slot'),
                    uid,
                    u.get('name',''), 
                    u.get('job',''), 
                    u.get('country',''),
                    (u.get('bio') or u.get('description','')),
                    ocean(u,'O'), 
                    ocean(u,'C'), 
                    ocean(u,'E'), 
                    ocean(u,'A'), 
                    ocean(u,'N')
                ])
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_users]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        safe_project = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        filename = f"Users_All_Personas_{run_id}_{safe_project}.xlsx"
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Cache-Control': 'no-store',
            }
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail='Excel export requires openpyxl library')
    except Exception as e:
        raise HTTPException(status_code=500, detail=f'Excel export failed: {str(e)}')

# _severity_for_category now imported from .utils


# _ingest_run_artifacts is imported from parent package (already imported above)


def resolve_project_run_dir(project_id_or_name: Optional[str]) -> pathlib.Path:
    print("Project Id received: {0}".format(project_id_or_name))
    if not project_id_or_name:
        raise HTTPException(status_code=400, detail='projectId required')
    items = read_projects()
    for it in items:
        if it.get('id') == project_id_or_name or it.get('name') == project_id_or_name:
            rid = it.get('run_id') or it.get('id')
            p = RUNS / rid
            if p.exists():
                return p
    raise HTTPException(status_code=404, detail='project not found')


async def preprocess_job(run_id: str, page: str, figma_url: str, verbose: bool, project_db_id: Optional[str] = None) -> None:
    run_dir = RUNS / run_id
    status_path = run_dir / 'status.json'
    log_path = run_dir / 'api_preprocess.log'
    status = {
        'run_id': run_id,
        'type': 'preprocess',
        'status': 'INPROGRESS',
        'page': page,
        'figma_url': figma_url,
        'started_at': time.time(),
        'updated_at': time.time(),
        'log': f"/runs-files/{run_id}/api_preprocess.log",
    }
    write_json(status_path, status)

    cmd = [
        PYTHON, str(ROOT / 'scripts' / 'run_one_step_extraction.py'),
        '--page', page,
        '--figma-url', figma_url,
        '--out-dir', run_id,
    ]
    if verbose:
        cmd.append('--verbose')
    env = os.environ.copy()
    env['PYTHONUNBUFFERED'] = '1'
    with open(log_path, 'ab') as lf:
        rc = subprocess.run(cmd, cwd=str(ROOT), env=env, stdout=lf, stderr=subprocess.STDOUT).returncode
    status['updated_at'] = time.time()
    status['finished_at'] = time.time()
    status['exit_code'] = rc
    status['status'] = 'COMPLETED' if rc == 0 else 'FAILED'
    write_json(status_path, status)
    # Mark project status and attach preprocess log URL
    if use_supabase_db():
        try:
            client = get_supabase()
            if project_db_id:
                client.table('projects').update({
                    'status': ('COMPLETED' if rc == 0 else 'FAILED')
                }).eq('id', project_db_id).execute()
        except Exception:
            pass
        try:
            project_name = 'project'
            if project_db_id:
                try:
                    res = get_supabase().table('projects').select('name').eq('id', project_db_id).limit(1).execute()
                    if res.data:
                        project_name = res.data[0].get('name') or project_name
                except Exception:
                    pass
            public_url = upload_log_to_supabase(log_path, project_name, 'preprocess')
            if project_db_id and public_url:
                get_supabase().table('projects').update({'meta': {'log_url': public_url}}).eq('id', project_db_id).execute()
        except Exception:
            pass
    else:
        try:
            if project_db_id:
                await execute('update projects set status=$1, updated_at=now() where id=$2',
                              'COMPLETED' if rc == 0 else 'FAILED', project_db_id)
        except Exception:
            pass
        # Upload log to Supabase (path: <project>/preprocess/<file>)
        try:
            # get project name from DB if available
            project_name = 'project'
            if project_db_id:
                row = await fetchrow('select name from projects where id=$1', project_db_id)
                if row:
                    project_name = row['name']
            public_url = upload_log_to_supabase(log_path, project_name, 'preprocess')
            if project_db_id and public_url:
                await execute('update projects set meta = coalesce(meta, \"{}\"::jsonb) || jsonb_build_object(\"log_url\", $1) where id=$2', public_url, project_db_id)
        except Exception:
            pass


async def tests_job(
    run_dir_str: str,
    goal: str,
    max_minutes: int,
    source_id: Optional[int] = None,
    target_id: Optional[int] = None,
    source_image: Optional[str] = None,
    target_image: Optional[str] = None,
    db_run_id: Optional[str] = None,
) -> None:
    print("Executing job...  Here is run_dir_str: {0}".format(run_dir_str))
    run_dir = pathlib.Path(run_dir_str)
    if not run_dir.is_absolute():
        run_dir = RUNS / run_dir_str
    run_id = run_dir.name
    status_path = run_dir / 'tests_status.json'
    log_path = run_dir / 'api_tests.log'
    status = {
        'run_id': run_id,
        'type': 'tests',
        'status': 'INPROGRESS',
        'goal': goal,
        'started_at': time.time(),
        'updated_at': time.time(),
        'log': f"/runs-files/{run_id}/api_tests.log",
    }

    print("Here is status: {0}".format(status))
    write_json(status_path, status)

    print("Now almost starting...  Here is cmd")

    try:
        # If a resolved persona plan exists, pass it to the runner via --persona-plan
        persona_plan_path = run_dir / 'tests' / 'persona_plan.json'
        cmd = [
            PYTHON, str(ROOT / 'scripts' / 'run_persona_inplace.py'),
            '--run-dir', str(run_dir),
            '--goal', goal,
            '--max-minutes', str(max_minutes or 2),
        ]
        if persona_plan_path.exists():
            cmd += ['--persona-plan', str(persona_plan_path)]
        if source_image and target_image:
            cmd += ['--source-image', source_image, '--target-image', target_image]
        else:
            cmd += ['--source-id', str(source_id), '--target-id', str(target_id)]

        env = os.environ.copy()
        env['PYTHONUNBUFFERED'] = '1'
        with open(log_path, 'ab') as lf:
            rc = subprocess.run(cmd, cwd=str(ROOT), env=env, stdout=lf, stderr=subprocess.STDOUT).returncode
        print("RC status...  Here is rc: {0}".format(rc))
        status['updated_at'] = time.time()
        status['finished_at'] = time.time()
        status['exit_code'] = rc
        status['status'] = 'COMPLETED' if rc == 0 else 'FAILED'
        write_json(status_path, status)
    except Exception as e:
        print("The test run executon exceptio error: {0}".format(e))

    # Ingest metrics BEFORE marking the DB run status as COMPLETED
    try:
        await _ingest_run_artifacts(run_dir, db_run_id)
    except Exception as e:
        print("New run ingestion exception error: {0}".format(e))
    if use_supabase_db():
        try:
            if db_run_id:
                # Try to upload tests.zip and persist URL in meta.report_url
                tests_zip_url = None
                try:
                    # Derive project name for storage path
                    project_name = None
                    try:
                        rr = get_supabase().table('runs').select('project_id').eq('id', db_run_id).limit(1).execute()
                        pid = rr.data[0]['project_id'] if rr.data else None
                        if pid:
                            pr = get_supabase().table('projects').select('name').eq('id', pid).limit(1).execute()
                            if pr.data:
                                project_name = pr.data[0]['name']
                    except Exception:
                        project_name = None
                    if project_name:
                        tests_zip_url = upload_tests_dir_zip(run_dir, project_name, run_id)
                except Exception:
                    tests_zip_url = None

                payload = {
                    'status': ('COMPLETED' if rc == 0 else 'FAILED'),
                    'finished_at': None,  # server timestamp handled by DB default/trigger if any
                    'log_path': str(status['log'])
                }
                if tests_zip_url:
                    payload['meta'] = {'report_url': tests_zip_url}
                get_supabase().table('runs').update(payload).eq('id', db_run_id).execute()
        except Exception as e:
            print("DB test failed update exception error: {0}".format(e))
    else:
        try:
            if db_run_id:
                await execute('update runs set status=$1, finished_at=now(), log_path=$2 where id=$3',
                              'COMPLETED' if rc == 0 else 'FAILED', str(status['log']), db_run_id)
        except Exception as e:
            print("DB test completed update exception error: {0}".format(e))
    # Upload tests log and artifacts to Supabase: <project>/runs/<test_run_id>/
    try:
        # find project name by joining runs->projects if possible
        project_name = None
        if db_run_id:
            if use_supabase_db():
                try:
                    rr = get_supabase().table('runs').select('project_id').eq('id', db_run_id).limit(1).execute()
                    pid = rr.data[0]['project_id'] if rr.data else None
                    if pid:
                        pr = get_supabase().table('projects').select('name').eq('id', pid).limit(1).execute()
                        if pr.data:
                            project_name = pr.data[0]['name']
                except Exception:
                    project_name = None
            else:
                row = await fetchrow('select p.name from runs r join projects p on r.project_id=p.id where r.id=$1', db_run_id)
                if row:
                    project_name = row['name']
        if project_name:
            upload_log_to_supabase(log_path, project_name, 'tests', run_id)
            # Upload all artifacts (images, csv, logs) under the run directory
            try:
                upload_run_artifacts(run_dir, project_name, run_id)
            except Exception as e:
                print("Upload run artifacts exception: {0}".format(e))
    except Exception as e:
        print("Upload run artifacts exception: {0}".format(e))

    


@router.post('/runs/preprocess')
async def start_preprocess(req: PreprocessReq, authorization: Optional[str] = Header(None)):
    RUNS.mkdir(parents=True, exist_ok=True)

    # Immediate check for valid user before scheduling work
    email = get_current_user(authorization)
    if not email:
        raise HTTPException(status_code=401, detail="unauthorized")
    # create a project id from provided name or page
    safe_name = (req.project_name or req.page).strip()
    safe_name = re.sub(r"[^a-zA-Z0-9_-]+", "_", safe_name)
    base_id = safe_name.lower()[:40] or 'project'
    timestamp = time.strftime('%Y%m%d_%H%M%S')
    rand = ''.join(random.choices(string.ascii_lowercase + string.digits, k=4))
    project_id = f"{base_id}_{timestamp}_{rand}"

    # backward compatibility for out_dir
    run_id = req.out_dir or project_id
    run_dir = RUNS / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    # candidate DB project_id (to return immediately)
    candidate_db_project_id = str(uuid.uuid4())

    log_path = run_dir / "api_preprocess.log"

    def log_message(msg: str, level: str = "INFO"):
        ts = time.strftime('%Y-%m-%d %H:%M:%S')
        line = f"[{ts}] {level}: {msg}"
        print(line)
        try:
            with open(log_path, "a") as f:
                f.write(line + "\n")
        except Exception as e:
            print("Failed to write to api_preprocess.log:", e)

    # mark INPROGRESS immediately (filesystem)
    write_json(run_dir / 'status.json', {
        'run_id': run_id,
        'type': 'preprocess',
        'status': 'INPROGRESS',
        'page': req.page,
        'figma_url': req.figma_url,
        'started_at': time.time(),
        'updated_at': time.time(),
        'log': f"/runs-files/{run_id}/api_preprocess.log",
    })
    log_message(f"Run {run_id} started. Project candidate {candidate_db_project_id}. Status set to INPROGRESS.")

    # helper for failures
    def fail_process(error_msg: str):
        log_message(error_msg, level="ERROR")
        try:
            write_json(run_dir / 'status.json', {
                'run_id': run_id,
                'type': 'preprocess',
                'status': 'FAILED',
                'error': error_msg,
                'updated_at': time.time(),
            })
            log_message("Status updated to FAILED in status.json", level="ERROR")
        except Exception as e:
            log_message(f"Failed to mark run as FAILED: {e}", level="ERROR")

    # ---- background async worker ----
    async def _deferred_work():
        log_message("Background worker started.")
        log_message(f"Authorization header = {authorization}")

        owner_email = get_current_user(authorization)
        log_message(f"Extracted owner_email={owner_email}")

        owner_id = None
        if owner_email:
            try:
                if use_supabase_db():
                    log_message("Fetching owner_id from Supabase...")
                    r = get_supabase().table('users').select('id').eq('email', owner_email).limit(1).execute()
                    if r.data:
                        owner_id = str(r.data[0]['id'])
                        log_message(f"Found owner_id={owner_id} in Supabase.")
                else:
                    log_message("Fetching owner_id from Postgres...")
                    row = await fetchrow('select id from users where email=$1', owner_email)
                    if row:
                        owner_id = str(row['id'])
                        log_message(f"Found owner_id={owner_id} in Postgres.")
            except Exception as e:
                log_message(f"Exception fetching owner_id: {e}", level="ERROR")
                owner_id = None

        # 🚨 Fail if no owner_id
        if not owner_id:
            fail_process("Missing owner_id (no valid user found)")
            return

        db_project_id: Optional[str] = None
        try:
            if use_supabase_db():
                log_message("Inserting project row into Supabase...")
                print("run_id: {0}".format(run_dir))
                res = get_supabase().table('projects').insert({
                    'id': candidate_db_project_id,
                    'owner_id': owner_id,
                    'name': req.project_name or req.page,
                    'figma_url': req.figma_url,
                    'figma_page': req.page,
                    'status': 'INPROGRESS',
                    'kind': 'figma',
                    'run_dir': str(run_dir)
                }, returning="representation").execute()  # ✅ FIXED: no .select() in Python client

                if res.data:
                    db_project_id = str(res.data[0]['id'])
                    log_message(f"Project inserted into Supabase with id={db_project_id}.")
            else:
                log_message("Inserting project row into Postgres...")
                row = await fetchrow(
                    'insert into projects (id, owner_id, name, figma_url, figma_page, status, kind) '
                    'values ($1,$2,$3,$4,$5,$6,$7) returning id',
                    candidate_db_project_id, owner_id,
                    req.project_name or req.page, req.figma_url, req.page, 'INITIATED', 'figma'
                )
                db_project_id = str(row['id']) if row else None
                if db_project_id:
                    log_message(f"Project inserted into Postgres with id={db_project_id}. Updating to INPROGRESS...")
                    await execute(
                        'update projects set status=$1, updated_at=now() where id=$2',
                        'INPROGRESS', db_project_id
                    )
                    log_message("Project status updated to INPROGRESS in Postgres.")
        except Exception as e:
            fail_process(f"DB insert failed: {e}")
            return

        # 🚨 Fail if DB insert did not return project id
        if not db_project_id:
            fail_process("DB insert failed: no project_id returned")
            return

        # Legacy JSON projects list
        try:
            log_message("Updating legacy projects.json...")
            items = read_projects()
            items.insert(0, {
                'id': project_id,
                'name': req.project_name or req.page,
                'run_id': run_id,
                'created_at': time.time(),
            })
            write_projects(items[:200])
            log_message("Legacy projects.json updated.")
        except Exception as e:
            log_message(f"Legacy JSON update failed: {e}", level="WARNING")

        # launch extraction
        log_message("Launching preprocess_job...")
        try:
            await preprocess_job(run_id, req.page, req.figma_url, req.verbose, db_project_id)
            log_message("preprocess_job completed successfully.")
        except Exception as e:
            fail_process(f"preprocess_job failed: {e}")

    # ✅ schedule background task without blocking
    import os as _os
    if _os.getenv('DISABLE_BACKGROUND', '').lower() in ('1','true','yes','y'):
        log_message("Background task scheduling skipped (DISABLE_BACKGROUND)")
    else:
        asyncio.create_task(_deferred_work())
    log_message("Background task scheduled.")

    # immediate response
    return {
        'accepted': True,
        'run_id': run_id,
        'project_id': project_id,
        'run_dir': str(run_dir),
        'status_url': f"/runs/{run_id}/status",
        'log': f"/runs-files/{run_id}/api_preprocess.log",
        'db': {'project_id': candidate_db_project_id}
    }


@router.post('/runs/tests-by-images')
async def start_tests_by_images(
    bg: BackgroundTasks,
    runDir: Optional[str] = Form(None),
    projectId: Optional[str] = Form(None),
    goal: str = Form(...),
    maxMinutes: int = Form(2),
    source: UploadFile = File(...),
    target: UploadFile = File(...),
    personas: Optional[str] = Form(None),
    exclusiveUsers: Optional[str] = Form(None),
    authorization: Optional[str] = Header(None)
):

    print("Yes")
    email = get_current_user(authorization)
    if not email:
        raise HTTPException(status_code=401, detail="unauthorized")

    # Step 1: Get owner_id from users table
    owner_id = None
    try:
        res = get_supabase().table("users").select("id").eq("email", email).limit(1).execute()
        if res.data:
            owner_id = str(res.data[0]['id'])
    except Exception as e:
        print(f"Error fetching owner_id: {e}")
    if not projectId:
        raise HTTPException(status_code=400, detail="projectId required")

    print("Noo")

    # ✅ Step 1: Check projectId in DB (Supabase or Postgres)
    project_exists = False
    print("yyy")
    if use_supabase_db():
        print("KKKK")
        try:
            print("iiii")
            res = get_supabase().table("projects").select("id,owner_id,run_dir").eq("id", projectId).limit(1).execute()
            print("Uuuuu")
            if res.data:
                print("TYYYTYY")
                print("res.data: {0}".format(res.data))
                project_exists = True
        except Exception as e:
            print(f"Supabase error checking projectId: {e}")
    else:
        try:
            row = await fetchrow("select id from projects where id=$1", projectId)
            if row:
                project_exists = True
        except Exception as e:
            print(f"Postgres error checking projectId: {e}")

    if not project_exists:
        raise HTTPException(status_code=404, detail="project not found")
    
    print("Project exists: {0}".format(project_exists))
    if owner_id != res.data[0]['owner_id']:
        raise HTTPException(status_code=401, detail="The project in invalid for the user")

    print("res.data: {0}".format(res.data))
    print("Whyyy")
    if runDir:
        run_dir = pathlib.Path(runDir)
        if not run_dir.is_absolute():
            run_dir = RUNS / runDir
    else:
        print("hhhbhdhbhdbdhb")
        run_dir = res.data[0]['run_dir']
        run_dir = pathlib.Path(run_dir)
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail='run_dir not found')
    test_run_id = str(uuid.uuid4())

    # Store "uploaded" images under run folder
    print("Hello")
    img_dir = run_dir / 'uploads'
    img_dir.mkdir(parents=True, exist_ok=True)
    source_path = img_dir / 'source.png'
    target_path = img_dir / 'target.png'
    source_path.write_bytes(await source.read())
    target_path.write_bytes(await target.read())

    print("img_dir: {0}, source_path: {1}, target_path: {2}".format(img_dir, source_path, target_path))

    # Parse personas + exclusivity flag
    persona_plan: list[dict[str, Any]] = []
    allow_overlap = True
    try:
        if personas:
            import json as _json
            pl = _json.loads(personas)
            if isinstance(pl, list):
                normalized = []
                for i, p in enumerate(pl, start=1):
                    normalized.append({
                        'slot': int(i),
                        'persona_id': int(p.get('personaId') or 0) if isinstance(p, dict) and str(p.get('personaId') or '').isdigit() else None,
                        'name': (p.get('name') or '') if isinstance(p, dict) else '',
                        'traits': (p.get('traits') or '') if isinstance(p, dict) else '',
                        'users': int(p.get('users') or 1) if isinstance(p, dict) else 1,
                    })
                persona_plan = normalized
        if isinstance(exclusiveUsers, str) and exclusiveUsers != '':
            allow_overlap = (exclusiveUsers.strip().lower() not in ('1','true','yes','y'))
    except Exception as e:
        print(f"persona plan parse error: {e}")

    # Persist requested plan in run_dir
    try:
        tests_root = run_dir / 'tests'
        tests_root.mkdir(parents=True, exist_ok=True)
        write_json(tests_root / 'persona_plan.json', { 'personas': persona_plan, 'allow_overlap': allow_overlap })
    except Exception as e:
        print(f"Failed writing persona_plan.json: {e}")

    # Resolve personas to concrete user ids from static persona list (server-side planning)
    try:
        persona_json_path = ROOT / 'users' / 'users.json'
        resolved = resolve_personas(persona_plan, persona_json_path, allow_overlap=allow_overlap)
        write_json(tests_root / 'persona_resolution.json', resolved)
    except Exception as e:
        print(f"persona resolution error: {e}")

    # mark INPROGRESS
    write_json(run_dir / 'tests_status.json', {
        'run_id': run_dir.name,
        'type': 'tests',
        'status': 'INPROGRESS',
        'test_run_id': test_run_id,
        'goal': goal,
        'started_at': time.time(),
        'updated_at': time.time(),
        'log': f"/runs-files/{run_dir.name}/api_tests.log",
        'uploads': {
            'source': f"/runs-files/{run_dir.name}/uploads/source.png",
            'target': f"/runs-files/{run_dir.name}/uploads/target.png",
        },
        'meta': { 'persona_plan': persona_plan, 'allow_overlap': allow_overlap }
    })

    # DB run insert
    db_run_id = test_run_id
    if use_supabase_db():
        try:
            if projectId:
                r = get_supabase().table('projects').select('id').eq('id', projectId).limit(1).execute()
                print("Here is r: {0}".format(r))
                if r.data:
                    print("Here is r.data: {0}".format(r.data))
                    try:
                        rr = get_supabase().table('runs').insert({
                            'id': test_run_id,
                            'project_id': str(r.data[0]['id']),
                            'kind': 'tests',
                            'status': 'INPROGRESS',
                            'goal': goal,
                            'log_path': f"/runs-files/{run_dir.name}/api_tests.log",
                            'run_dir': str(run_dir),
                            'meta': {'uploads': True, 'persona_plan': persona_plan, 'allow_overlap': allow_overlap},
                        }, returning="representation").execute() 
                    except Exception as e:
                        print("Here is error: {0}".format(e))
                    print("Here is rr: {0}".format(rr))
                    print("Is it executed")
                    if rr.data:
                        print("I think it is executed")
                        db_run_id = str(rr.data[0]['id'])
                    else:
                        print("No no not")
        except Exception:
            pass
    else:
        try:
            row = await fetchrow('select id from projects where id=$1', projectId) if projectId else None
            if row:
                rr = await fetchrow('insert into runs (project_id, kind, status, goal, log_path, run_dir, meta) values ($1,$2,$3,$4,$5,$6,$7) returning id',
                                    str(row['id']), 'tests', 'INITIATED', goal, f"/runs-files/{run_dir.name}/api_tests.log", str(run_dir), json.dumps({'uploads': True}))
                db_run_id = str(rr['id']) if rr else None
                await execute('update runs set status=$1, started_at=now() where id=$2', 'INPROGRESS', db_run_id)
        except Exception:
            pass

    # Persist resolution to DB meta when available (best-effort)
    if use_supabase_db():
        try:
            res_obj = None
            try:
                res_obj = json.loads((tests_root / 'persona_resolution.json').read_text('utf-8'))
            except Exception:
                pass
            if res_obj is not None and db_run_id:
                get_supabase().table('runs').update({ 'meta': { 'uploads': True, 'persona_plan': persona_plan, 'allow_overlap': allow_overlap, 'persona_resolution': res_obj } }).eq('id', db_run_id).execute()
        except Exception as e:
            print(f"failed to persist persona_resolution to DB: {e}")

    #bg.add_task(tests_job, str(run_dir), goal, int(maxMinutes or 2), None, None, str(source_path), str(target_path), db_run_id)

    import os as _os
    if _os.getenv('DISABLE_BACKGROUND', '').lower() in ('1','true','yes','y'):
        log_message("tests_job scheduling skipped (DISABLE_BACKGROUND)")
    else:
        asyncio.create_task(
            tests_job(
                str(run_dir), goal, int(maxMinutes or 2),
                None, None, str(source_path), str(target_path), db_run_id
            )
        )

    return {
        'accepted': True,
        'run_dir': str(run_dir),
        'test_run_id': test_run_id,
        'uploads': {'source': f"/runs-files/{run_dir.name}/uploads/source.png", 'target': f"/runs-files/{run_dir.name}/uploads/target.png"},
        'status_url': f"/runs/{run_dir.name}/status",
        'log': f"/runs-files/{run_dir.name}/api_tests.log",
        'db': {'run_id': db_run_id}
    }


@router.get('/runs/{run_id}/logs.zip')
async def download_test_logs(
    run_id: str,
    project_id: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    """
    Download test run logs as a zip file.
    Creates a zip of the entire test_run_<timestamp> folder.
    """
    try:
        # Verify authentication if using Supabase
        if use_supabase_db() and authorization:
            from ..auth_utils import get_current_user
            try:
                token = authorization.replace('Bearer ', '') if authorization else None
                if token:
                    await get_current_user(token)
            except Exception as e:
                raise HTTPException(status_code=401, detail=f"Authentication failed: {str(e)}")

        # Find the run directory
        run_dir = RUNS / run_id
        if not run_dir.exists():
            raise HTTPException(status_code=404, detail=f"Run directory not found: {run_id}")

        # Create zip file in memory
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for file_path in run_dir.rglob('*'):
                if file_path.is_file():
                    try:
                        # Get relative path from run_dir
                        arc_name = file_path.relative_to(run_dir).as_posix()
                        zf.write(file_path, arc_name)
                    except Exception as e:
                        print(f"Error adding {file_path} to zip: {e}")
                        continue

        # Prepare response
        buf.seek(0)
        zip_data = buf.getvalue()
        
        return Response(
            content=zip_data,
            media_type='application/zip',
            headers={
                'Content-Disposition': f'attachment; filename="{run_id}_logs.zip"'
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"Error creating logs zip: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to create logs zip: {str(e)}")
